import pandas as pd
from rdflib import URIRef, BNode, Literal, Graph
from rdflib.namespace import RDF, RDFS, FOAF, XSD
from rdflib import Namespace
import numpy as np
import math
import sys
import argparse
import json
import html
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import openpyxl

'''
wb = openpyxl.Workbook()
ws = wb.active
# 2枚目のシートを作成
ws2 = wb.create_sheet("Sheet2")

# A1にSheet2という文字列を設定して、リンク先に同ドキュメント内の"Sheet2"シートのA1を指定
ws["A1"].value = "Sheet2"
ws["A1"].hyperlink = "test.xlsx#sheet2!A1"
# A2に"CNN"という文字列を設定して、リンク先にCNNのアドレスを指定
ws["A2"].value = "CNN"
ws["A2"].hyperlink = "https://edition.cnn.com/cnn10"
# A3に"はてな"という文字列を設定して、リンク先に「はてな」のアドレスを指定
ws.cell(row=3, column=1).value = u"はてな"
ws.cell(row=3, column=1).hyperlink = "http://www.hatena.ne.jp/"
# HYPERLINK関数自体を設定することも可能
ws.cell(row=4, column=1).value = '=HYPERLINK("https://www.amazon.co.jp/", "Amazon")'
ws.cell(row=4, column=1).font = openpyxl.styles.fonts.Font(color='FF0000')
# 保存
wb.save('test.xlsx')
'''

############

'''
from openpyxl import Workbook, load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active

r_count = ws.max_row
c_count = ws.max_column

for j in range(2, r_count):
    for i in range(0, c_count):
        cell = ws.cell(row=j+1, column=i+1)
        print(j, i, cell.value)
'''

############

with open("/Users/nakamurasatoru/git/d_sat/u-renja/static/iiif2/collection/top.json") as f:
    df2 = json.load(f)

manifests = df2["manifests"]

images = {}

for m in manifests:
    metadata = m["metadata"]

    num = -1
    identifier = ""

    for obj in metadata:
        if "Description" in obj["label"]:
            num = obj["value"][0].split("通番: ")[1]

            num = int(num)

        elif "identifier" in obj["label"]:
            identifier = obj["value"]

    if num not in images:
        images[num] = []

    images[num].append({
        "id": m["@id"],
        "identifier": identifier
    })

'''
for num in images:
    print(num, images[num])
'''

#############

url = "https://taishozo.github.io/db/iiif/kandomokuroku/manifest.json"

df = requests.get(url).json()

indexMap = {}

canvases = df["sequences"][0]["canvases"]
for i in range(len(canvases)):
    c = canvases[i]
    res = c["images"][0]["resource"]["@id"]
    if "p." in res:
        p = res.split("p.")[1].split(".")[0]
        indexMap[p] = i + 1

'''
for p in indexMap:
    print(p, indexMap[p], int(p) - int(indexMap[p]))
'''

hyperlink = Font(underline='single', color='0563C1')

def read_excel(path):
    # df = pd.read_excel(path, sheet_name=0, header=None, index_col=None, engine="openpyxl")

    wb = load_workbook(path)
    ws = wb.active

    r_count = ws.max_row
    c_count = ws.max_column

    map = {}

    for i in range(0, c_count):
        label = ws.cell(row=1+1, column=i+1).value
        map[i] = label

    for j in range(2, r_count):
        id = ws.cell(row=j+1, column=0+1).value

        if id == "" or id == None:
            continue

        経典番号 =  ws.cell(row=j+1, column=1+1).value
        枝番 = ws.cell(row=j+1, column=9+1).value
        if 枝番 == "" or 枝番 == None:
            枝番 = ""

        e1 = ws.cell(row=j+1, column=8+1).value if ws.cell(row=j+1, column=8+1).value != None else ws.cell(row=j+1, column=8+1).value # ""
        e2 = ws.cell(row=j+1, column=10+1).value if ws.cell(row=j+1, column=10+1).value != None else ws.cell(row=j+1, column=10+1).value # ""
        e3 = ws.cell(row=j+1, column=11+1).value if ws.cell(row=j+1, column=11+1).value != None else ws.cell(row=j+1, column=11+1).value # ""
        e4 = ws.cell(row=j+1, column=12+1).value if ws.cell(row=j+1, column=12+1).value != None else ws.cell(row=j+1, column=12+1).value # ""
        e5 = ws.cell(row=j+1, column=13+1).value if ws.cell(row=j+1, column=13+1).value != None else ws.cell(row=j+1, column=13+1).value # ""

        uri_sat = "https://21dzk.l.u-tokyo.ac.jp/SAT2018/"+e1+枝番+"_."+str(e2).zfill(2)+"."+str(e3).zfill(4)+e4+str(e5).zfill(2)+".html"

        title = ws.cell(row=j+1, column=3+1).value
        
        ws.cell(row=j+1, column=3+1).value = '=HYPERLINK("{}", "{}")'.format(uri_sat, title)
        ws.cell(row=j+1, column=3+1).font = hyperlink
        
        num1 = ws.cell(row=j+1, column=114+1).value #df.iloc[j, 114]
        if num1 != "" and num1 != None:
            ws.cell(row=j+1, column=114+1).value = '=HYPERLINK("https://taishozo.github.io/u-renja/search/?fc-通番={}", "{}")'.format(num1, num1)
            ws.cell(row=j+1, column=114+1).font = hyperlink
            
            # df.iloc[j, 114] = "[{}]({})".format(num1, num1)

        num2 = ws.cell(row=j+1, column=118+1).value #df.iloc[j, 114]
        if num2 != "" and num2 != None:
            ws.cell(row=j+1, column=118+1).value = '=HYPERLINK("https://taishozo.github.io/u-renja/search/?fc-通番={}", "{}")'.format(num2, num2)
            ws.cell(row=j+1, column=118+1).font = hyperlink

        kando = ws.cell(row=j+1, column=122+1).value #df.iloc[j, 114]
        if kando != "" and kando != None:
            pos = int(kando) - 152
            ws.cell(row=j+1, column=122+1).value = '=HYPERLINK("http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest=https://taishozo.github.io/db/iiif/kandomokuroku/manifest.json&pos={}", "{}")'.format(pos, int(kando))
            ws.cell(row=j+1, column=122+1).font = hyperlink

        folder1 = ws.cell(row=j+1, column=127+1).value #df.iloc[j, 114]
        if folder1 != "" and folder1 != None:

            x = ws.cell(row=j+1, column=128+1).value
            y = ws.cell(row=j+1, column=129+1).value

            print(folder1, x, y)

            uuid1 = folder1 + "_" + str(x).zfill(4) + "_" + str(y).zfill(4)

            num1 = int(num1)

            if num1 in images:

                value1 = ""
                    
                arr = images[num1]

                for a in arr:
                    if uuid1 == a["identifier"]:
                        value1 = a["id"]

                if value1 == "":
                    print("missing", uuid1, "value1", value1)
                else:
                    ws.cell(row=j+1, column=126+1).value = '=HYPERLINK("http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest={}", "{}")'.format(value1, ws.cell(row=j+1, column=126+1).value)
                    ws.cell(row=j+1, column=126+1).font = hyperlink
                    # df.iloc[j, 126] = "[{}]({})".format(df.iloc[j, 126], value1)

        folder2 = ws.cell(row=j+1, column=131+1).value #df.iloc[j, 114]
        if folder2 != "" and folder2 != None:

            x = ws.cell(row=j+1, column=132+1).value
            y = ws.cell(row=j+1, column=132+1).value

            print(folder2, x, y)

            uuid2 = folder2 + "_" + str(x).zfill(4) + "_" + str(y).zfill(4)

            num1 = int(num1)

            if num1 in images:

                value2 = ""
                    
                arr = images[num1]

                for a in arr:
                    if uuid2 == a["identifier"]:
                        value2 = a["id"]

                if value2 == "":
                    print("missing", uuid2, "value2", value2)
                else:
                    ws.cell(row=j+1, column=130+1).value = '=HYPERLINK("http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest={}", "{}")'.format(value2, ws.cell(row=j+1, column=130+1).value)
                    ws.cell(row=j+1, column=130+1).font = hyperlink

        '''
        num2 = df.iloc[j, 118]
        if not pd.isnull(num2):
            df.iloc[j, 118] = "[{}](https://taishozo.github.io/u-renja/search/?fc-通番={})".format(num2, num2)

        kando = df.iloc[j, 122]
        if not pd.isnull(kando):
            pos = int(kando) - 152
            df.iloc[j, 122] = "[{}](http://codh.rois.ac.jp/software/iiif-curation-viewer/demo/?manifest=https://taishozo.github.io/db/iiif/kandomokuroku/manifest.json&pos={})".format(int(kando), pos)

        folder1 = df.iloc[j, 127]
        if not pd.isnull(folder1):

            uuid1 = folder1 + "_" + str(df.iloc[j, 128]).zfill(4) + "_" + str(df.iloc[j, 129]).zfill(4)

            num1 = int(num1)

            if num1 in images:

                value1 = ""
                    
                arr = images[num1]

                for a in arr:
                    if uuid1 == a["identifier"]:
                        value1 = a["id"]

                if value1 == "":
                    print(uuid1, "value1", value1)
                else:
                    df.iloc[j, 126] = "[{}]({})".format(df.iloc[j, 126], value1)

            # dir1 = df.iloc[j, 127]

        folder2 = df.iloc[j, 131]
        if not pd.isnull(folder2):
            uuid2 = folder2 + "_" + str(df.iloc[j, 132]).zfill(4) + "_" + str(df.iloc[j, 133]).zfill(4)

            num1 = int(num1)

            if num1 in images:

                value2 = ""
                    
                arr = images[num1]

                for a in arr:
                    if uuid2 == a["identifier"]:
                        value2 = a["id"]

                if value2 == "":
                    print(uuid2, "value2", value2)
                else:
                    df.iloc[j, 130] = "[{}]({})".format(df.iloc[j, 130], value2)


        # print(num, kando, title1, dir1, title2, dir2)

        '''

    '''
    writer = pd.ExcelWriter('../static/metadata/data.xlsx', engine="openpyxl")
    df.to_excel(writer, index=False, header=False)

    #Excelファイルを保存
    writer.save()
    #Excelファイルを閉じる
    writer.close()
    '''

    wb.save('../static/metadata/data.xlsx')

path = "data/『大正新脩大蔵経』底本・校本一覧データベース20210622.xlsx"
# data1 = read_excel(path)
read_excel(path)